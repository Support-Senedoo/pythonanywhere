"""
Vérification de l'équilibre des pièces comptables Odoo via XML-RPC API.
Exporte les écritures déséquilibrées dans un fichier Excel avec détail des lignes.

Usage :
    python verifier_equilibre_pieces.py

Dépendances : openpyxl (pip install openpyxl)
"""

import xmlrpc.client
import sys
import io
from collections import defaultdict
from datetime import datetime

# Force UTF-8 sur la console Windows (évite UnicodeEncodeError avec les accents)
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# CONFIGURATION
# ──────────────────────────────────────────────
URL       = "https://lysa.odoo.com"
DB        = "lysa"
USERNAME  = "support@senedoo.com"
PASSWORD  = "2026@Senedoo"
DATE_FROM = "2025-08-05"
DATE_TO   = "2025-08-19"

OUTPUT_FILE = f"pieces_desequilibrees_{DATE_FROM}_{DATE_TO}.xlsx"

# Tolérance d'arrondi (en devise comptable)
TOLERANCE = 0.005

# ──────────────────────────────────────────────
# CONNEXION
# ──────────────────────────────────────────────

def connect():
    """Authentification XML-RPC et retour (uid, proxy models)."""
    print(f"Connexion à {URL} (base : {DB}) …")
    common = xmlrpc.client.ServerProxy(f"{URL}/xmlrpc/2/common", allow_none=True)
    version = common.version()
    print(f"  Odoo {version.get('server_version', '?')}")

    uid = common.authenticate(DB, USERNAME, PASSWORD, {})
    if not uid:
        sys.exit("ERREUR : authentification échouée. Vérifiez les identifiants.")
    print(f"  Authentifié — UID {uid}\n")

    models = xmlrpc.client.ServerProxy(f"{URL}/xmlrpc/2/object", allow_none=True)
    return uid, models


def call(models, uid, model, method, args, kwargs=None):
    """Raccourci execute_kw."""
    return models.execute_kw(DB, uid, PASSWORD, model, method, args, kwargs or {})

# ──────────────────────────────────────────────
# RÉCUPÉRATION DES DONNÉES
# ──────────────────────────────────────────────

def fetch_moves(uid, models):
    """Récupère toutes les pièces comptables validées dans la période."""
    domain = [
        ("date",  ">=", DATE_FROM),
        ("date",  "<=", DATE_TO),
        ("state", "=",  "posted"),
    ]
    move_ids = call(models, uid, "account.move", "search", [domain])
    print(f"Pièces (state=posted) trouvées : {len(move_ids)}")

    if not move_ids:
        return []

    moves = call(models, uid, "account.move", "read", [move_ids], {
        "fields": ["name", "date", "ref", "journal_id", "move_type"],
    })
    return moves


def fetch_lines_for_moves(uid, models, move_ids):
    """
    Récupère TOUTES les lignes comptables des pièces en un seul appel
    (évite N+1 requêtes).
    Retourne un dict { move_id: [lines] }.
    """
    domain = [("move_id", "in", move_ids)]
    # Batch : 1 000 lignes à la fois pour éviter les timeouts
    BATCH = 1000
    all_lines = []
    offset = 0
    while True:
        chunk = call(models, uid, "account.move.line", "search_read", [domain], {
            "fields": [
                "move_id", "account_id", "name",
                "debit", "credit",
                "partner_id", "amount_currency", "currency_id",
            ],
            "limit":  BATCH,
            "offset": offset,
        })
        all_lines.extend(chunk)
        if len(chunk) < BATCH:
            break
        offset += BATCH

    print(f"Lignes comptables récupérées : {len(all_lines)}")

    # Regroupement par move_id
    by_move = defaultdict(list)
    for line in all_lines:
        by_move[line["move_id"][0]].append(line)
    return by_move

# ──────────────────────────────────────────────
# VÉRIFICATION DE L'ÉQUILIBRE
# ──────────────────────────────────────────────

def check_balance(moves, lines_by_move):
    """Retourne la liste des pièces déséquilibrées avec métadonnées."""
    unbalanced = []
    for move in moves:
        mid   = move["id"]
        lines = lines_by_move.get(mid, [])
        total_debit  = sum(l["debit"]  for l in lines)
        total_credit = sum(l["credit"] for l in lines)
        diff = abs(total_debit - total_credit)
        if diff > TOLERANCE:
            unbalanced.append({
                "move":         move,
                "lines":        lines,
                "total_debit":  total_debit,
                "total_credit": total_credit,
                "diff":         diff,
            })
    return unbalanced

# ──────────────────────────────────────────────
# EXPORT EXCEL
# ──────────────────────────────────────────────

# Palette de couleurs
C_HEADER_DARK  = "1F4E79"   # bleu foncé
C_HEADER_MID   = "2E75B6"   # bleu moyen
C_WHITE        = "FFFFFF"
C_RED_LIGHT    = "FFDCE0"
C_RED_TEXT     = "C00000"
C_GREY_ROW     = "F2F2F2"
C_TOTAL_FILL   = "FFF2CC"   # jaune clair pour totaux


def _hdr(font_color=C_WHITE, bg=C_HEADER_DARK, bold=True, center=True):
    return {
        "font":  Font(bold=bold, color=font_color),
        "fill":  PatternFill(start_color=bg, end_color=bg, fill_type="solid"),
        "align": Alignment(horizontal="center", vertical="center", wrap_text=True),
    }


def _apply(cell, font=None, fill=None, align=None, number_format=None):
    if font:          cell.font      = font
    if fill:          cell.fill      = fill
    if align:         cell.alignment = align
    if number_format: cell.number_format = number_format


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_row(ws, row, values):
    for col, val in enumerate(values, 1):
        ws.cell(row=row, column=col, value=val)


def build_excel(unbalanced, output_path):
    wb = openpyxl.Workbook()

    # ── Feuille 1 : Résumé ──────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Résumé"
    ws1.freeze_panes = "A2"

    # Titre
    ws1.merge_cells("A1:H1")
    title_cell = ws1["A1"]
    title_cell.value = (
        f"Pièces déséquilibrées — {DATE_FROM} au {DATE_TO}  "
        f"({len(unbalanced)} pièce{'s' if len(unbalanced) != 1 else ''})"
    )
    _apply(title_cell,
           font=Font(bold=True, size=13, color=C_WHITE),
           fill=PatternFill(start_color=C_HEADER_DARK, end_color=C_HEADER_DARK, fill_type="solid"),
           align=Alignment(horizontal="center", vertical="center"))
    ws1.row_dimensions[1].height = 28

    # En-têtes
    hdrs1 = ["N° Pièce", "Date", "Référence", "Journal", "Type",
             "Total Débit", "Total Crédit", "Écart (abs.)"]
    for col, h in enumerate(hdrs1, 1):
        c = ws1.cell(row=2, column=col, value=h)
        _apply(c, **_hdr(bg=C_HEADER_MID))
    ws1.row_dimensions[2].height = 22

    # Données
    num_fmt = '#,##0.00'
    for r, item in enumerate(unbalanced, 3):
        m  = item["move"]
        bg = PatternFill(start_color=C_GREY_ROW, end_color=C_GREY_ROW, fill_type="solid") \
             if r % 2 == 0 else None

        row_vals = [
            m["name"],
            m["date"],
            m["ref"] or "",
            m["journal_id"][1] if m["journal_id"] else "",
            m["move_type"],
            round(item["total_debit"],  2),
            round(item["total_credit"], 2),
            round(item["diff"],         2),
        ]
        _write_row(ws1, r, row_vals)
        for col in range(1, 9):
            c = ws1.cell(row=r, column=col)
            if bg:
                c.fill = bg
            if col >= 6:
                c.number_format = num_fmt
        # Écart en rouge
        ecart_cell = ws1.cell(row=r, column=8)
        _apply(ecart_cell,
               font=Font(bold=True, color=C_RED_TEXT),
               fill=PatternFill(start_color=C_RED_LIGHT, end_color=C_RED_LIGHT, fill_type="solid"))

    _set_col_widths(ws1, [20, 12, 25, 22, 16, 14, 14, 14])

    # ── Feuille 2 : Détail des lignes ───────────────────────────────────────
    ws2 = wb.create_sheet("Détail des lignes")
    ws2.freeze_panes = "A2"

    hdrs2 = ["N° Pièce", "Date", "Journal", "ID Ligne", "Compte",
             "Libellé ligne", "Partenaire", "Débit", "Crédit",
             "Devise", "Mt. Devise"]
    for col, h in enumerate(hdrs2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        _apply(c, **_hdr(bg=C_HEADER_DARK))
    ws2.row_dimensions[1].height = 22

    detail_row = 2
    for item in unbalanced:
        m    = item["move"]
        move_name = m["name"]
        move_date = m["date"]
        journal   = m["journal_id"][1] if m["journal_id"] else ""

        for line in item["lines"]:
            row_vals = [
                move_name,
                move_date,
                journal,
                line["id"],
                line["account_id"][1]  if line["account_id"]  else "",
                line["name"]           or "",
                line["partner_id"][1]  if line["partner_id"]  else "",
                line["debit"],
                line["credit"],
                line["currency_id"][1] if line["currency_id"] else "",
                line["amount_currency"],
            ]
            _write_row(ws2, detail_row, row_vals)
            for col in [8, 9, 11]:
                ws2.cell(row=detail_row, column=col).number_format = num_fmt
            detail_row += 1

        # Ligne de total par pièce
        total_fill = PatternFill(start_color=C_TOTAL_FILL, end_color=C_TOTAL_FILL, fill_type="solid")
        for col in range(1, 12):
            ws2.cell(row=detail_row, column=col).fill = total_fill

        ws2.cell(row=detail_row, column=6).value = "TOTAL PIECE"
        ws2.cell(row=detail_row, column=6).font  = Font(bold=True)
        for col, val in [(8, item["total_debit"]), (9, item["total_credit"])]:
            c = ws2.cell(row=detail_row, column=col, value=round(val, 2))
            c.number_format = num_fmt
            c.font = Font(bold=True)

        ecart = ws2.cell(row=detail_row, column=10,
                         value=f"ÉCART : {round(item['diff'], 2)}")
        _apply(ecart,
               font=Font(bold=True, color=C_RED_TEXT),
               fill=PatternFill(start_color=C_RED_LIGHT, end_color=C_RED_LIGHT, fill_type="solid"))

        detail_row += 2   # ligne blanche entre les pièces

    _set_col_widths(ws2, [20, 12, 20, 10, 30, 35, 28, 13, 13, 10, 13])

    # ── Feuille 3 : Aucun déséquilibre (message si OK) ──────────────────────
    if not unbalanced:
        ws3 = wb.create_sheet("Résultat")
        ws3["A1"].value = "Toutes les pièces comptables sont équilibrées."
        ws3["A1"].font  = Font(bold=True, color="375623", size=12)

    # Sauvegarde
    wb.save(output_path)
    print(f"\nFichier Excel exporté : {output_path}")

# ──────────────────────────────────────────────
# POINT D'ENTRÉE
# ──────────────────────────────────────────────

def main():
    print("=" * 58)
    print("  VÉRIFICATION ÉQUILIBRE PIÈCES COMPTABLES — ODOO API")
    print("=" * 58)
    print(f"  Periode  : {DATE_FROM} au {DATE_TO}")
    print(f"  Base     : {DB} ({URL})")
    print(f"  Tolérance: {TOLERANCE} €")
    print("=" * 58 + "\n")

    uid, models = connect()

    moves = fetch_moves(uid, models)
    if not moves:
        print("Aucune pièce validée dans cette période. Fin.")
        return

    move_ids     = [m["id"] for m in moves]
    lines_by_move = fetch_lines_for_moves(uid, models, move_ids)

    print("\nVérification de l'équilibre…")
    unbalanced = check_balance(moves, lines_by_move)

    print(f"\n{'-'*40}")
    if unbalanced:
        print(f"  /!\\ {len(unbalanced)} piece(s) DESEQUILIBREE(s) detectee(s) !")
        for item in unbalanced:
            m = item["move"]
            print(f"     * {m['name']}  ({m['date']})  "
                  f"débit={item['total_debit']:.2f}  "
                  f"crédit={item['total_credit']:.2f}  "
                  f"écart={item['diff']:.4f}")
    else:
        print("  OK - Toutes les pieces sont equilibrees.")
    print(f"{'─'*40}\n")

    build_excel(unbalanced, OUTPUT_FILE)
    print("Terminé.")


if __name__ == "__main__":
    main()
